# ef_agent.py
from __future__ import annotations
import os, json, re, io, csv
from typing import Dict, Any, List, Optional

import ollama
import pypdf, docx2txt
from openpyxl import load_workbook
from langchain_community.vectorstores import FAISS
from langchain_huggingface import HuggingFaceEmbeddings

BASE_URL   = os.getenv("OPENAI_BASE_URL", "http://localhost:11434").rstrip("/")
DEFAULT_LLM = os.getenv("EF_AI_MODEL", "deepseek-r1:8b")
INDEX_DIR  = os.getenv("INDEX_DIR", "faiss_index_EF")
UPLOAD_DIR = os.getenv("UPLOAD_DIR", "uploads")
MEM_PATH   = os.path.join("outputs", "memory.json")

_EMB = HuggingFaceEmbeddings(model_name="sentence-transformers/paraphrase-multilingual-MiniLM-L12-v2")

# -------------------- Memoria --------------------
def _load_memory() -> Dict[str, Any]:
    try:
        with open(MEM_PATH, "r", encoding="utf-8") as f: return json.load(f)
    except Exception: return {"facts": [], "lang": None}
def _save_memory(mem: Dict[str, Any]) -> None:
    os.makedirs(os.path.dirname(MEM_PATH), exist_ok=True)
    with open(MEM_PATH, "w", encoding="utf-8") as f: json.dump(mem, f, ensure_ascii=False, indent=2)

# -------------------- Ollama helpers --------------------
def _client() -> ollama.Client: return ollama.Client(host=BASE_URL)
def _chat_json(model: str, messages: List[Dict[str,str]], max_tokens=160) -> Optional[dict]:
    resp = _client().chat(
        model=model, format="json", messages=messages,
        options={"temperature":0.1, "num_ctx":4096, "num_predict":max_tokens}
    )
    txt = resp.get("message",{}).get("content","").strip()
    try: return json.loads(txt)
    except Exception:
        m = re.search(r"\{.*\}", txt, re.S)
        return json.loads(m.group(0)) if m else None
def _chat_text(model: str, messages: List[Dict[str,str]], max_tokens=360) -> str:
    resp = _client().chat(
        model=model, messages=messages,
        options={"temperature":0.2, "num_ctx":8192, "num_predict":max_tokens}
    )
    return resp.get("message",{}).get("content","").strip()
def _pretty(d: Any, maxlen=1600) -> str:
    s = json.dumps(d, ensure_ascii=False, indent=2)
    return s if len(s)<=maxlen else s[:maxlen]+"…"

# -------------------- Lectura de archivos --------------------
def _read_pdf(path: str) -> str:
    with open(path, "rb") as f:
        reader = pypdf.PdfReader(f)
        return "\n".join(pg.extract_text() or "" for pg in reader.pages)
def _read_docx(path: str) -> str: return docx2txt.process(path) or ""
def _read_txt(path: str) -> str:
    with open(path, "r", encoding="utf-8", errors="ignore") as f: return f.read()
def _read_csv(path: str) -> str:
    out = []
    with open(path, "r", encoding="utf-8", errors="ignore") as f:
        for row in csv.reader(f): out.append(" | ".join(row))
    return "\n".join(out)
def _read_xlsx(path: str) -> str:
    wb = load_workbook(path, data_only=True)
    chunks = []
    for name in wb.sheetnames:
        ws = wb[name]
        rows = []
        for r in ws.iter_rows(values_only=True):
            rows.append(" | ".join("" if v is None else str(v) for v in r))
        chunks.append(f"[{name}]\n" + "\n".join(rows))
    return "\n\n".join(chunks)

# -------------------- Tools --------------------
def tool_list_uploads() -> Dict[str, Any]:
    files = []
    if os.path.isdir(UPLOAD_DIR):
        for name in sorted(os.listdir(UPLOAD_DIR)):
            p = os.path.join(UPLOAD_DIR, name)
            if os.path.isfile(p): files.append(name)
    return {"files": files}

def tool_read_pdf(filename: str, max_chars: int = 8000) -> Dict[str, Any]:
    path = os.path.join(UPLOAD_DIR, filename)
    if not os.path.exists(path): return {"error": f"Archivo no encontrado: {filename}"}
    text = re.sub(r"\s+", " ", _read_pdf(path)).strip()
    return {"text": (text[:max_chars]+"…") if len(text)>max_chars else text, "filename": filename}
def tool_read_docx(filename: str, max_chars: int = 8000) -> Dict[str, Any]:
    path = os.path.join(UPLOAD_DIR, filename)
    if not os.path.exists(path): return {"error": f"Archivo no encontrado: {filename}"}
    text = re.sub(r"\s+", " ", _read_docx(path)).strip()
    return {"text": (text[:max_chars]+"…") if len(text)>max_chars else text, "filename": filename}
def tool_read_txt(filename: str, max_chars: int = 12000) -> Dict[str, Any]:
    path = os.path.join(UPLOAD_DIR, filename)
    if not os.path.exists(path): return {"error": f"Archivo no encontrado: {filename}"}
    text = re.sub(r"\s+", " ", _read_txt(path)).strip()
    return {"text": (text[:max_chars]+"…") if len(text)>max_chars else text, "filename": filename}
def tool_read_csv(filename: str, max_chars: int = 12000) -> Dict[str, Any]:
    path = os.path.join(UPLOAD_DIR, filename)
    if not os.path.exists(path): return {"error": f"Archivo no encontrado: {filename}"}
    text = re.sub(r"\s+", " ", _read_csv(path)).strip()
    return {"text": (text[:max_chars]+"…") if len(text)>max_chars else text, "filename": filename}
def tool_read_xlsx(filename: str, max_chars: int = 12000) -> Dict[str, Any]:
    path = os.path.join(UPLOAD_DIR, filename)
    if not os.path.exists(path): return {"error": f"Archivo no encontrado: {filename}"}
    text = re.sub(r"\s+", " ", _read_xlsx(path)).strip()
    return {"text": (text[:max_chars]+"…") if len(text)>max_chars else text, "filename": filename}
def tool_read_file(filename: Optional[str] = None, max_chars: int = 8000) -> Dict[str, Any]:
    if not filename: return {"error": "Falta 'filename'. Llama antes a list_uploads o proporciona un nombre."}
    low = filename.lower()
    if low.endswith(".pdf"):  return tool_read_pdf(filename, max_chars=max_chars)
    if low.endswith((".docx",".doc")): return tool_read_docx(filename, max_chars=max_chars)
    if low.endswith((".txt",".md")):   return tool_read_txt(filename, max_chars=max_chars)
    if low.endswith(".csv"):  return tool_read_csv(filename, max_chars=max_chars)
    if low.endswith(".xlsx"): return tool_read_xlsx(filename, max_chars=max_chars)
    return {"error": f"Extensión no soportada para '{filename}'."}

def _load_index() -> Optional[FAISS]:
    if not os.path.isdir(INDEX_DIR): return None
    return FAISS.load_local(INDEX_DIR, _EMB, allow_dangerous_deserialization=True)
def tool_search_index(query: str, k: int = 5, unit: Optional[str]=None) -> Dict[str, Any]:
    db = _load_index()
    if db is None: return {"error": "No hay índice FAISS. Usa el indexador primero."}
    res = db.similarity_search_with_score(query, k=24)
    out = []
    for d, sc in res:
        if unit and d.metadata.get("unidad") != unit: continue
        out.append({
            "snippet": (d.page_content or "").replace("\n"," ")[:240],
            "source": d.metadata.get("source","?"),
            "unidad": d.metadata.get("unidad","?"),
            "score": float(sc)
        })
        if len(out) >= k: break
    if not out:
        for d, sc in res[:k]:
            out.append({
                "snippet": (d.page_content or "").replace("\n"," ")[:240],
                "source": d.metadata.get("source","?"),
                "unidad": d.metadata.get("unidad","?"),
                "score": float(sc)
            })
    return {"results": out}

def tool_summarize_text(text: str, sentences: int = 5, model: str = DEFAULT_LLM) -> Dict[str, Any]:
    sys = f"Resume el texto en exactamente {sentences} frases claras en español."
    out = _chat_text(model, [{"role":"system","content":sys},{"role":"user","content":text}], max_tokens=240)
    return {"summary": out}

def tool_remember(fact: str) -> Dict[str, Any]:
    mem = _load_memory()
    if fact and fact not in mem["facts"]:
        mem["facts"].append(fact); _save_memory(mem)
    return {"facts": mem["facts"]}
def tool_recall() -> Dict[str, Any]: return _load_memory()

def tool_correct_with_rubric(rubric_file: str, work_file: str, model: str = DEFAULT_LLM) -> Dict[str, Any]:
    rpath = os.path.join(UPLOAD_DIR, rubric_file)
    if not os.path.exists(rpath): return {"error": f"Rúbrica no encontrada: {rubric_file}"}
    if rubric_file.lower().endswith(".xlsx"): rubric_text = _read_xlsx(rpath)
    elif rubric_file.lower().endswith(".csv"): rubric_text = _read_csv(rpath)
    elif rubric_file.lower().endswith((".yaml",".yml",".txt",".md")): rubric_text = _read_txt(rpath)
    else: return {"error":"Formato de rúbrica no soportado (XLSX/CSV/YAML/TXT)."}
    wpath = os.path.join(UPLOAD_DIR, work_file)
    if not os.path.exists(wpath): return {"error": f"Trabajo no encontrado: {work_file}"}
    if work_file.lower().endswith(".pdf"): work_text = _read_pdf(wpath)
    elif work_file.lower().endswith(".docx"): work_text = _read_docx(wpath)
    else: return {"error":"Formato de trabajo no soportado (DOCX o PDF)."}
    sys = (
      "Eres corrector. A partir de la RÚBRICA (texto) y del TRABAJO, devuelve SOLO JSON con: "
      "{\"overall_score\":1|2|3|4, \"criteria\":[{\"name\":\"...\",\"score\":1|2|3|4,\"reasons\":\"<=40 palabras\"}], "
      "\"feedback_student\":\"<=120 palabras\"}."
    )
    user = f"RÚBRICA:\n{rubric_text}\n\nTRABAJO:\n{work_text[:8000]}"
    raw = _chat_text(model, [{"role":"system","content":sys},{"role":"user","content":user}], max_tokens=400)
    try:
        m = re.search(r"\{.*\}", raw, re.S)
        return json.loads(m.group(0)) if m else {"error":"No se pudo extraer JSON de evaluación."}
    except Exception:
        return {"error":"JSON inválido devuelto por el modelo."}

TOOLS: Dict[str, Any] = {
    "list_uploads":        {"fn": tool_list_uploads,        "desc":"Lista archivos en /uploads"},
    "read_pdf":            {"fn": tool_read_pdf,            "desc":"Lee PDF"},
    "read_docx":           {"fn": tool_read_docx,           "desc":"Lee DOCX"},
    "read_txt":            {"fn": tool_read_txt,            "desc":"Lee TXT/MD"},
    "read_csv":            {"fn": tool_read_csv,            "desc":"Lee CSV"},
    "read_xlsx":           {"fn": tool_read_xlsx,           "desc":"Lee XLSX"},
    "read_file":           {"fn": tool_read_file,           "desc":"Lee cualquier archivo por extensión"},
    "search_index":        {"fn": tool_search_index,        "desc":"Busca en FAISS"},
    "summarize_text":      {"fn": tool_summarize_text,      "desc":"Resumen breve"},
    "remember":            {"fn": tool_remember,            "desc":"Guardar hecho en memoria"},
    "recall":              {"fn": tool_recall,              "desc":"Mostrar memoria"},
    "correct_with_rubric": {"fn": tool_correct_with_rubric, "desc":"Corrección con rúbrica"}
}
def tools_manifest() -> str:
    return "\n".join([f"- {k}: {v['desc']}" for k,v in TOOLS.items()])

# -------- alias + heurística --------
TOOL_ALIASES = {
    "list_files":"list_uploads","listar_archivos":"list_uploads","ls":"list_uploads","files":"list_uploads",
    "read":"read_file","readfile":"read_file","read_any":"read_file","read_document":"read_file",
    "open":"read_file","open_file":"read_file","leer":"read_file",
    "readpdf":"read_pdf","readdocx":"read_docx","readtxt":"read_txt","readcsv":"read_csv","readxlsx":"read_xlsx",
    "search":"search_index","buscar":"search_index","rag":"search_index","retrieve":"search_index",
    "remember_fact":"remember","memorize":"remember","guardar":"remember",
    "recall_memory":"recall","memoria":"recall",
    "summarize":"summarize_text","resumen":"summarize_text","resume":"summarize_text",
    "correct":"correct_with_rubric","rubric":"correct_with_rubric","rúbrica":"correct_with_rubric","corregir":"correct_with_rubric"
}
def _normalize_tool(name: str) -> str:
    key = (name or "").strip().lower()
    if key in TOOLS: return key
    if key in TOOL_ALIASES: return TOOL_ALIASES[key]
    # heurística
    if any(w in key for w in ["list","listar","files","ls"]): return "list_uploads"
    if any(w in key for w in ["read","leer","open"]): return "read_file"
    if any(w in key for w in ["search","buscar","rag","retriev"]): return "search_index"
    if any(w in key for w in ["summar","resum"]): return "summarize_text"
    if "remember" in key or "guardar" in key: return "remember"
    if "recall" in key or "memoria" in key: return "recall"
    if any(w in key for w in ["rubric","rúbric","correg"]): return "correct_with_rubric"
    return key

# -------------------- System prompt --------------------
SYS = (
  "¿Prefieres trabajar en español o inglés?\n"
  "Eres Agente EF-AI con herramientas. MODOS: AGENTE_GENERAL, CHAT_RAG, RUBRICA.\n"
  "REGLAS: Responde SIEMPRE en JSON válido para acciones (una sola línea). Para terminar usa {\"final\":\"…\"}. "
  "No muestres <think>.\n"
  "HERRAMIENTAS:\n" + tools_manifest() + "\n"
  "PROCEDIMIENTO:\n"
  "1) Si dudas del modo, pide 1 aclaración (como final). 2) Si necesitas archivos: {\"tool\":\"list_uploads\",\"args\":{}} y luego una lectura. "
  "3) Puedes usar {\"tool\":\"read_file\",\"args\":{\"filename\":\"NOMBRE\"}}.\n"
  "EJEMPLO:\n"
  "User: Lista los archivos y lee el primero. Resúmelo en 5 frases.\n"
  "Assistant: {\"tool\":\"list_uploads\",\"args\":{}}\n"
  "User: (resultado)\n"
  "Assistant: {\"tool\":\"read_file\",\"args\":{\"filename\":\"EL_PRIMERO.pdf\"}}\n"
  "User: (texto)\n"
  "Assistant: {\"final\":\"Resumen en 5 frases…\"}\n"
)

# -------------------- Controlador --------------------
def _pick_tool_name(obj: dict) -> Optional[str]:
    # acepta varias claves posibles
    for k in ("tool","action","call","function","name"):
        if k in obj and isinstance(obj[k], str):
            return obj[k]
        if k in obj and isinstance(obj[k], dict) and "name" in obj[k]:
            return obj[k]["name"]
    return None

def run_agent(goal: str, model: str = DEFAULT_LLM, max_steps: int = 12) -> Dict[str, Any]:
    messages: List[Dict[str,str]] = [{"role":"system","content":SYS}]
    mem = _load_memory()
    if mem.get("facts"): messages.append({"role":"user","content":"Memoria:\n" + "\n".join(mem["facts"])})
    messages.append({"role":"user","content":f"OBJETIVO:\n{goal}"})

    transcript: List[Dict[str, Any]] = []
    last_text_blob: Optional[str] = None
    last_files: List[str] = []

    for _ in range(max_steps):
        raw = _chat_json(model, messages, max_tokens=140)
        # algunos modelos devuelven lista de acciones
        obj = raw[0] if isinstance(raw, list) and raw else raw
        if not isinstance(obj, dict):
            messages.append({"role":"user","content":"Responde SOLO JSON válido (una línea). Si puedes cerrar, usa {\"final\":\"…\"}."})
            continue

        if "final" in obj:
            transcript.append({"final": obj["final"]})
            return {"ok": True, "steps": transcript, "final": obj["final"]}

        tool_raw = _pick_tool_name(obj)
        args = (obj.get("args") or obj.get("arguments") or {}) or {}

        tool = _normalize_tool(tool_raw or "")
        # autocompletar filename si falta
        if tool in ("read_file","read_pdf","read_docx","read_txt","read_csv","read_xlsx"):
            if not args.get("filename") and last_files:
                args["filename"] = last_files[0]

        if tool not in TOOLS:
            # No abortar: elige heurística por el objetivo si podemos
            goal_low = (goal or "").lower()
            if any(w in goal_low for w in ["lista","listar","files","archiv"]): tool = "list_uploads"
            elif any(w in goal_low for w in ["lee","leer","read","open"]): tool = "read_file"
            elif any(w in goal_low for w in ["buscar","search","rag"]): tool = "search_index"
            elif any(w in goal_low for w in ["resumen","resume","summar"]): tool = "summarize_text"
            elif any(w in goal_low for w in ["rúbric","rubric","correg"]): tool = "correct_with_rubric"

        if tool not in TOOLS:
            transcript.append({"error":"TOOL_UNKNOWN", "obj": obj, "goal": goal})
            # intenta cerrar con lo que tengamos
            if last_text_blob:
                summary = tool_summarize_text(last_text_blob, sentences=5, model=model)["summary"]
                transcript.append({"final": summary, "note":"failsafe_summary"})
                return {"ok": True, "steps": transcript, "final": summary}
            return {"ok": False, "error":"TOOL_UNKNOWN", "steps": transcript}

        # ejecutar herramienta
        try: result = TOOLS[tool]["fn"](**args)
        except TypeError as e: result = {"error": f"Argumentos inválidos: {e}"}
        except Exception as e: result = {"error": f"Fallo ejecutando {tool}: {e}"}

        transcript.append({"tool": tool, "args": args, "result": result})

        # persistencia auxiliar
        if tool == "list_uploads" and isinstance(result, dict):
            last_files = result.get("files") or []
        if isinstance(result, dict):
            if "text" in result: last_text_blob = result["text"]
            elif tool == "search_index" and result.get("results"):
                last_text_blob = "\n".join(r.get("snippet","") for r in result["results"])

        obs = _pretty({"tool": tool, "result": result})
        messages.append({"role":"user","content":f"OBSERVACION:\n{obs}\n\nSi ya puedes terminar, responde con {{\"final\":\"…\"}}; si no, emite otra llamada de herramienta."})

    if last_text_blob:
        summary = tool_summarize_text(last_text_blob, sentences=5, model=model)["summary"]
        transcript.append({"final": summary, "note":"failsafe_summary"})
        return {"ok": True, "steps": transcript, "final": summary}

    return {"ok": False, "error":"MAX_STEPS_REACHED", "steps": transcript}